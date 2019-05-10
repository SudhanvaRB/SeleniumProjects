'''
Created on Nov 29, 2018

@author: subharad
'''

import openpyxl
from openpyxl.styles import Alignment

def getRowCount(file,sheetName):
    wk = openpyxl.load_workbook(file)
    #sht = wk.get_sheet_by_name(sheetName)
    sht = wk[sheetName]
    return (sht.max_row)

def getColCount(file,sheetName):
    wk = openpyxl.load_workbook(file)
    sht = wk[sheetName]
    return (sht.max_column)

def readData(file,sheetName,rowno,colno):
    wk = openpyxl.load_workbook(file)
    sht = wk[sheetName]
    return sht.cell(row=rowno, column=colno).value

def getCellvalue(file,sheetName,cellval):
    wk = openpyxl.load_workbook(file)
    sht = wk[sheetName]
    return sht[cellval].value

def writeData(file,sheetname,rownum,colno,data):
    wk = openpyxl.load_workbook(file)
    sheet = wk[sheetname]
    sheet.cell(row=rownum, column=colno).value=data
    sheet.cell(row=rownum, column=colno).alignment = Alignment(wrap_text=True)
    wk.save(file)



